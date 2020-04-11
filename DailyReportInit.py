from openpyxl import load_workbook
import xlrd

wb = load_workbook("Daily report.xlsx")
sheets = wb.sheetnames
Sheet1 = wb[sheets[0]]

excel_location = "Daily report.xlsx"
workbook_read = xlrd.open_workbook(excel_location)
worksheet_read = workbook_read.sheet_by_index(0)

jobs_col = 1
goal_status_col = 19
cycle_progress_col = 23
spend_col = 25
apps_progress_col = 26
handler_col = 44
detail_col = 45

dic_details = {}

for row in range(1, worksheet_read.nrows):
    jobs = worksheet_read.cell_value(row, jobs_col)
    Goal_status = worksheet_read.cell_value(row, goal_status_col)
    cycle_progress = worksheet_read.cell_value(row, cycle_progress_col) * 100
    spend = worksheet_read.cell_value(row, spend_col) * 100
    apps_progress = worksheet_read.cell_value(row, apps_progress_col) * 100

    if Goal_status == "Paused":
        dic_details[row] = "Inactive"
    elif spend > 100:
        dic_details[row] = "Over spend"
    elif Goal_status == "Goal Achieved":
        dic_details[row] = "Goal Achieved"
    elif cycle_progress >= apps_progress + 15 and cycle_progress >= spend + 15:
        dic_details[row] = "Need to create new variants or increase CPC"
    elif cycle_progress <= apps_progress + 15 and cycle_progress <= spend + 5 and spend <= apps_progress + 7:
        dic_details[row] = "Good performance"
    elif cycle_progress <= 4:
        dic_details[row] = "New cycle"
    elif jobs == 1:
        dic_details[row] = "Need to create variants"

for detail in dic_details:
    Sheet1.cell(row=detail + 1, column=detail_col).value = dic_details[detail]
    Sheet1.cell(row=detail + 1, column=handler_col).value = "Ailon"

wb.save("Daily report.xlsx")


